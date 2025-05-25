from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework import status
from rest_framework.permissions import IsAuthenticated
from rest_framework_simplejwt.authentication import JWTAuthentication
from openpyxl import load_workbook
from django.core.exceptions import ObjectDoesNotExist
from link_db.models import Creneau, ChargesEnseignement, Enseignants, Formations
from zipfile import BadZipFile
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator  


@method_decorator(csrf_exempt, name='dispatch')
class UploadExcel_creneau(APIView):
    parser_classes = (MultiPartParser, FormParser)
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('excel_file')

        if not excel_file:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        # Log file details for debugging
        print(f"Received file: {excel_file.name}")
        print(f"File size: {excel_file.size} bytes")
        print(f"Content type: {excel_file.content_type}")

        try:
            # Check file extension
            file_name = excel_file.name.lower()
            
            if file_name.endswith('.csv'):
                # Handle CSV files
                import csv
                import io
                from datetime import datetime, date, time
                
                # Read CSV file
                decoded_file = excel_file.read().decode('utf-8-sig')  # utf-8-sig handles BOM
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string)
                
                rows = list(reader)
                if len(rows) < 2:
                    return Response({
                        'error': 'CSV file appears to be empty or only contains headers'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                inserted = []
                skipped = []
                
                # Skip header row
                for row_num, row in enumerate(rows[1:], start=2):
                    if len(row) < 3:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Insufficient columns. Expected at least 3, got {len(row)}'
                        })
                        continue
                    
                    # Process row data
                    date_str = row[0].strip() if row[0] else None
                    heure_str = row[1].strip() if row[1] else None
                    salle = row[2].strip() if row[2] else None
                    
                    if not (date_str and heure_str):
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': 'Missing date or time'
                        })
                        continue
                    
                    # Parse date and time
                    try:
                        # Parse date (expecting YYYY-MM-DD format)
                        date_creneau = datetime.strptime(date_str, '%Y-%m-%d').date()
                    except ValueError:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Invalid date format: {date_str}. Expected YYYY-MM-DD'
                        })
                        continue
                    
                    try:
                        # Parse time (expecting HH:MM:SS or HH:MM format)
                        if ':' in heure_str:
                            time_parts = heure_str.split(':')
                            if len(time_parts) == 2:
                                heure_creneau = datetime.strptime(heure_str, '%H:%M').time()
                            else:
                                heure_creneau = datetime.strptime(heure_str, '%H:%M:%S').time()
                        else:
                            skipped.append({
                                'row': f'Row {row_num}',
                                'reason': f'Invalid time format: {heure_str}. Expected HH:MM:SS or HH:MM'
                            })
                            continue
                    except ValueError:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Invalid time format: {heure_str}. Expected HH:MM:SS or HH:MM'
                        })
                        continue

                    # Check for duplicates
                    if Creneau.objects.filter(date_creneau=date_creneau, heure_creneau=heure_creneau, salle=salle).exists():
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Duplicate entry: {date_str} {heure_str} {salle}'
                        })
                        continue

                    try:
                        creneau = Creneau(date_creneau=date_creneau, heure_creneau=heure_creneau, salle=salle)
                        creneau.save()
                        inserted.append({
                            'date_creneau': date_creneau.isoformat(),
                            'heure_creneau': heure_creneau.strftime('%H:%M:%S'),
                            'salle': salle
                        })
                    except Exception as e:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Database error: {str(e)}'
                        })
                
            else:
                # Handle .xlsx files with openpyxl
                from datetime import datetime, date, time
                
                workbook = load_workbook(filename=excel_file, data_only=True)
                sheet = workbook.active

                inserted = []
                skipped = []

                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not row or len(row) < 3:
                        continue
                        
                    date_val, heure_val, salle = row[0], row[1], row[2]

                    if not (date_val and heure_val):
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': 'Missing date or time'
                        })
                        continue

                    # Handle date conversion
                    if isinstance(date_val, datetime):
                        date_creneau = date_val.date()
                    elif isinstance(date_val, date):
                        date_creneau = date_val
                    elif isinstance(date_val, str):
                        try:
                            date_creneau = datetime.strptime(date_val, '%Y-%m-%d').date()
                        except ValueError:
                            skipped.append({
                                'row': f'Row {row_num}',
                                'reason': f'Invalid date format: {date_val}'
                            })
                            continue
                    else:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Invalid date type: {type(date_val)}'
                        })
                        continue

                    # Handle time conversion
                    if isinstance(heure_val, time):
                        heure_creneau = heure_val
                    elif isinstance(heure_val, datetime):
                        heure_creneau = heure_val.time()
                    elif isinstance(heure_val, str):
                        try:
                            if ':' in heure_val:
                                time_parts = heure_val.split(':')
                                if len(time_parts) == 2:
                                    heure_creneau = datetime.strptime(heure_val, '%H:%M').time()
                                else:
                                    heure_creneau = datetime.strptime(heure_val, '%H:%M:%S').time()
                            else:
                                skipped.append({
                                    'row': f'Row {row_num}',
                                    'reason': f'Invalid time format: {heure_val}'
                                })
                                continue
                        except ValueError:
                            skipped.append({
                                'row': f'Row {row_num}',
                                'reason': f'Invalid time format: {heure_val}'
                            })
                            continue
                    else:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Invalid time type: {type(heure_val)}'
                        })
                        continue

                    # Ensure salle is a string
                    salle = str(salle).strip() if salle else None

                    # Check for duplicates
                    if Creneau.objects.filter(date_creneau=date_creneau, heure_creneau=heure_creneau, salle=salle).exists():
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Duplicate entry'
                        })
                        continue

                    try:
                        creneau = Creneau(date_creneau=date_creneau, heure_creneau=heure_creneau, salle=salle)
                        creneau.save()
                        inserted.append({
                            'date_creneau': date_creneau.isoformat(),
                            'heure_creneau': heure_creneau.strftime('%H:%M:%S'),
                            'salle': salle
                        })
                    except Exception as e:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Database error: {str(e)}'
                        })

            # If nothing was processed, the file might be in wrong format
            if len(inserted) == 0 and len(skipped) == 0:
                return Response({
                    'error': 'No valid data found in the file. Please check the file format.',
                    'expected_columns': ['date_creneau', 'heure_creneau', 'salle'],
                    'expected_formats': {
                        'date_creneau': 'YYYY-MM-DD',
                        'heure_creneau': 'HH:MM:SS or HH:MM',
                        'salle': 'text'
                    },
                    'note': 'First row should be headers, data starts from row 2'
                }, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'message': f'{len(inserted)} new time slots inserted, {len(skipped)} skipped.',
                'inserted': inserted,
                'skipped': skipped,
                'summary': {
                    'total_rows': len(inserted) + len(skipped),
                    'inserted_count': len(inserted),
                    'skipped_count': len(skipped)
                }
            }, status=status.HTTP_201_CREATED)

        except BadZipFile:
            return Response({
                'error': 'Invalid Excel file. The file appears to be corrupted or is not a valid Excel file.',
                'hint': 'Please ensure you are uploading a valid .xlsx or .csv file.'
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({
                'error': f'Error processing file: {str(e)}',
                'type': type(e).__name__,
                'hint': 'Please ensure your file has the correct format and column structure'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

@method_decorator(csrf_exempt, name='dispatch')
class ChargesEnseignement_xlsx(APIView):
    parser_classes = (MultiPartParser, FormParser)
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            workbook = load_workbook(filename=excel_file, data_only=True)
            sheet = workbook.active

            inserted = []
            skipped = []

            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):  # skip header
                # Skip empty rows
                if not row or all(cell is None for cell in row):
                    continue
                    
                # Check if row has enough columns
                if len(row) < 11:
                    skipped.append({
                        'row': f'Row {row_num}',
                        'reason': f'Insufficient columns. Expected at least 11, got {len(row)}'
                    })
                    continue

                # Adapt indices to your Excel file structure
                palier = row[0] if row[0] else None
                specialite = row[1] if row[1] else None
                semestre = row[2] if row[2] else None
                section = row[3] if row[3] else None
                groupe = row[4] if row[4] else None
                type_charge = row[5] if row[5] else None
                intitule_module = row[6] if row[6] else None
                abv_module = row[7] if row[7] else None
                code_enseignant_val = row[8] if row[8] else None
                annee_universitaire = row[9] if row[9] else None
                formation_val = row[10] if len(row) > 10 and row[10] else None

                # Validation: Required field
                if not annee_universitaire:
                    skipped.append({'row': row, 'reason': 'Missing annee_universitaire'})
                    continue

                # Attempt to get Enseignant instance or None
                enseignant = None
                if code_enseignant_val:
                    try:
                        enseignant = Enseignants.objects.get(Code_Enseignant=code_enseignant_val)
                    except ObjectDoesNotExist:
                        skipped.append({'row': row, 'reason': f'Enseignant with code {code_enseignant_val} not found'})
                        continue

                # Attempt to get Formation instance or None
                formation = None
                if formation_val:
                    try:
                        formation = Formations.objects.get(pk=formation_val)
                    except ObjectDoesNotExist:
                        skipped.append({'row': row, 'reason': f'Formation with id {formation_val} not found'})
                        continue

                # Prevent duplicates: check if this charge exists by unique fields
                exists = ChargesEnseignement.objects.filter(
                    palier=palier,
                    specialite=specialite,
                    semestre=semestre,
                    section=section,
                    groupe=groupe,
                    type=type_charge,
                    intitulé_module=intitule_module,
                    abv_module=abv_module,
                    Code_Enseignant_id=enseignant,
                    annee_universitaire=annee_universitaire,
                    formation=formation
                ).exists()

                if exists:
                    skipped.append({'row': row, 'reason': 'Duplicate entry'})
                    continue

                try:
                    # Create and save
                    charge = ChargesEnseignement(
                        palier=palier,
                        specialite=specialite,
                        semestre=semestre,
                        section=section,
                        groupe=groupe,
                        type=type_charge,
                        intitulé_module=intitule_module,
                        abv_module=abv_module,
                        Code_Enseignant_id=enseignant,  # Note: Django expects the model instance here
                        annee_universitaire=annee_universitaire,
                        formation=formation
                    )
                    charge.save()
                    inserted.append({
                        'palier': palier,
                        'specialite': specialite,
                        'semestre': semestre,
                        'section': section,
                        'groupe': groupe,
                        'type': type_charge,
                        'intitule_module': intitule_module,
                        'abv_module': abv_module,
                        'code_enseignant': code_enseignant_val,
                        'annee_universitaire': annee_universitaire,
                        'formation': formation_val
                    })
                except Exception as e:
                    skipped.append({
                        'row': f'Row {row_num}',
                        'reason': f'Database error: {str(e)}'
                    })

            return Response({
                'message': f'{len(inserted)} rows inserted, {len(skipped)} rows skipped.',
                'inserted': inserted,
                'skipped': skipped,
            }, status=status.HTTP_201_CREATED)

        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@method_decorator(csrf_exempt, name='dispatch')
class UploadEnseignants_xlsx(APIView):
    parser_classes = (MultiPartParser, FormParser)
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        # Log file details for debugging
        print(f"Received file: {excel_file.name}")
        print(f"File size: {excel_file.size} bytes")
        print(f"Content type: {excel_file.content_type}")

        try:
            # Check file extension
            file_name = excel_file.name.lower()
            
            if file_name.endswith('.csv'):
                # Handle CSV files
                import csv
                import io
                
                # Read CSV file
                decoded_file = excel_file.read().decode('utf-8-sig')  # utf-8-sig handles BOM
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string)
                
                rows = list(reader)
                if len(rows) < 2:
                    return Response({
                        'error': 'CSV file appears to be empty or only contains headers'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                inserted = []
                skipped = []
                
                # Skip header row
                for row_num, row in enumerate(rows[1:], start=2):
                    if len(row) < 12:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Insufficient columns. Expected at least 12, got {len(row)}'
                        })
                        continue
                    
                    # Process row data (same logic as Excel processing)
                    code_enseignant = row[0].strip() if row[0] else None
                    nom = row[1].strip() if len(row) > 1 else ''
                    prenom = row[2].strip() if len(row) > 2 else ''
                    nom_jeune_fille = row[3].strip() if len(row) > 3 else ''
                    genre = row[4].strip() if len(row) > 4 else ''
                    etat = row[5].strip() if len(row) > 5 else ''
                    departement = row[6].strip() if len(row) > 6 else ''
                    grade = row[7].strip() if len(row) > 7 else ''
                    email1 = row[8].strip() if len(row) > 8 else ''
                    email2 = row[9].strip() if len(row) > 9 else ''
                    tel1 = row[10].strip() if len(row) > 10 else ''
                    tel2 = row[11].strip() if len(row) > 11 else ''
                    
                    # Validate and save (same as Excel logic)
                    if not code_enseignant:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': 'Missing Code_Enseignant (column 1)'
                        })
                        continue

                    if Enseignants.objects.filter(Code_Enseignant=code_enseignant).exists():
                        existing = Enseignants.objects.get(Code_Enseignant=code_enseignant)
                        skipped.append({
                            'row': f'Row {row_num} (Code: {code_enseignant})',
                            'reason': f'Teacher already exists: {existing.nom} {existing.prenom}'
                        })
                        continue

                    try:
                        enseignant = Enseignants(
                            Code_Enseignant=code_enseignant,
                            nom=nom,
                            prenom=prenom,
                            nom_jeune_fille=nom_jeune_fille,
                            genre=genre,
                            etat=etat,
                            département=departement,
                            grade=grade,
                            email1=email1,
                            email2=email2,
                            tel1=tel1,
                            tel2=tel2
                        )
                        enseignant.save()
                        
                        inserted.append({
                            'Code_Enseignant': code_enseignant,
                            'nom': enseignant.nom,
                            'prenom': enseignant.prenom,
                            'departement': enseignant.département
                        })
                    except Exception as e:
                        skipped.append({
                            'row': f'Row {row_num} (Code: {code_enseignant})',
                            'reason': f'Database error: {str(e)}'
                        })
                
            elif file_name.endswith('.xls'):
                # Handle old Excel format
                import xlrd
                
                # Save uploaded file temporarily
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix='.xls') as tmp_file:
                    for chunk in excel_file.chunks():
                        tmp_file.write(chunk)
                    tmp_file_path = tmp_file.name
                
                try:
                    # Open with xlrd
                    workbook = xlrd.open_workbook(tmp_file_path)
                    sheet = workbook.sheet_by_index(0)
                    
                    if sheet.nrows < 2:
                        return Response({
                            'error': 'Excel file appears to be empty or only contains headers'
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    inserted = []
                    skipped = []
                    
                    for row_num in range(1, sheet.nrows):
                        row = sheet.row_values(row_num)
                        
                        if len(row) < 12:
                            skipped.append({
                                'row': f'Row {row_num + 1}',
                                'reason': f'Insufficient columns. Expected at least 12, got {len(row)}'
                            })
                            continue
                        
                        # Process row (same logic as CSV)
                        code_enseignant = str(row[0]).strip() if row[0] else None
                        nom = str(row[1]).strip() if len(row) > 1 else ''
                        prenom = str(row[2]).strip() if len(row) > 2 else ''
                        nom_jeune_fille = str(row[3]).strip() if len(row) > 3 else ''
                        genre = str(row[4]).strip() if len(row) > 4 else ''
                        etat = str(row[5]).strip() if len(row) > 5 else ''
                        departement = str(row[6]).strip() if len(row) > 6 else ''
                        grade = str(row[7]).strip() if len(row) > 7 else ''
                        email1 = str(row[8]).strip() if len(row) > 8 else ''
                        email2 = str(row[9]).strip() if len(row) > 9 else ''
                        tel1 = str(row[10]).strip() if len(row) > 10 else ''
                        tel2 = str(row[11]).strip() if len(row) > 11 else ''
                        
                        # Remove .0 from numeric codes if present
                        if code_enseignant and code_enseignant.endswith('.0'):
                            code_enseignant = code_enseignant[:-2]
                        
                        # Validate and save (same as above)
                        if not code_enseignant:
                            skipped.append({
                                'row': f'Row {row_num + 1}',
                                'reason': 'Missing Code_Enseignant (column 1)'
                            })
                            continue

                        if Enseignants.objects.filter(Code_Enseignant=code_enseignant).exists():
                            existing = Enseignants.objects.get(Code_Enseignant=code_enseignant)
                            skipped.append({
                                'row': f'Row {row_num + 1} (Code: {code_enseignant})',
                                'reason': f'Teacher already exists: {existing.nom} {existing.prenom}'
                            })
                            continue

                        try:
                            enseignant = Enseignants(
                                Code_Enseignant=code_enseignant,
                                nom=nom,
                                prenom=prenom,
                                nom_jeune_fille=nom_jeune_fille,
                                genre=genre,
                                etat=etat,
                                département=departement,
                                grade=grade,
                                email1=email1,
                                email2=email2,
                                tel1=tel1,
                                tel2=tel2
                            )
                            enseignant.save()
                            
                            inserted.append({
                                'Code_Enseignant': code_enseignant,
                                'nom': enseignant.nom,
                                'prenom': enseignant.prenom,
                                'departement': enseignant.département
                            })
                        except Exception as e:
                            skipped.append({
                                'row': f'Row {row_num + 1} (Code: {code_enseignant})',
                                'reason': f'Database error: {str(e)}'
                            })
                finally:
                    # Clean up temp file
                    import os
                    os.unlink(tmp_file_path)
                    
            else:
                # Handle .xlsx files with openpyxl
                workbook = load_workbook(filename=excel_file, data_only=True)
                sheet = workbook.active

                inserted = []
                skipped = []
                
                # Check if sheet has any data
                if sheet.max_row < 2:
                    return Response({
                        'error': 'Excel file appears to be empty or only contains headers'
                    }, status=status.HTTP_400_BAD_REQUEST)

                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    # Skip empty rows
                    if not row or all(cell is None for cell in row):
                        continue
                        
                    # Check if row has at least 12 columns (matching Teacher interface)
                    if len(row) < 12:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Insufficient columns. Expected at least 12, got {len(row)}'
                        })
                        continue

                    # Extract values based on Teacher interface field order
                    code_enseignant = row[0] if row[0] else None
                    nom = row[1] if row[1] else ''
                    prenom = row[2] if row[2] else ''
                    nom_jeune_fille = row[3] if row[3] else ''
                    genre = row[4] if row[4] else ''
                    etat = row[5] if row[5] else ''  # Maps to 'status' in frontend
                    departement = row[6] if row[6] else ''
                    grade = row[7] if row[7] else ''
                    email1 = row[8] if row[8] else ''
                    email2 = row[9] if row[9] else ''
                    tel1 = row[10] if row[10] else ''
                    tel2 = row[11] if row[11] else ''

                    # Validate required field
                    if not code_enseignant:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': 'Missing Code_Enseignant (column 1)'
                        })
                        continue

                    # Convert code_enseignant to string and strip whitespace
                    code_enseignant = str(code_enseignant).strip()

                    # Check for duplicates - note the capital C in Code_Enseignant
                    if Enseignants.objects.filter(Code_Enseignant=code_enseignant).exists():
                        existing = Enseignants.objects.get(Code_Enseignant=code_enseignant)
                        skipped.append({
                            'row': f'Row {row_num} (Code: {code_enseignant})',
                            'reason': f'Teacher already exists: {existing.nom} {existing.prenom}'
                        })
                        continue

                    try:
                        # Create teacher - only with fields that exist in the model
                        enseignant = Enseignants(
                            Code_Enseignant=code_enseignant,  # Note: capital C
                            nom=str(nom).strip(),
                            prenom=str(prenom).strip(),
                            nom_jeune_fille=str(nom_jeune_fille).strip(),
                            genre=str(genre).strip(),
                            etat=str(etat).strip(),
                            département=str(departement).strip(),  # Note: DB uses 'département' with accent
                            grade=str(grade).strip(),
                            email1=str(email1).strip(),
                            email2=str(email2).strip(),
                            tel1=str(tel1).strip(),
                            tel2=str(tel2).strip()
                        )
                        
                        # Save to database
                        enseignant.save()
                        
                        inserted.append({
                            'Code_Enseignant': code_enseignant,
                            'nom': enseignant.nom,
                            'prenom': enseignant.prenom,
                            'departement': enseignant.département
                        })
                        
                    except Exception as e:
                        skipped.append({
                            'row': f'Row {row_num} (Code: {code_enseignant})',
                            'reason': f'Database error: {str(e)}'
                        })

            # If nothing was processed, the file might be in wrong format
            if len(inserted) == 0 and len(skipped) == 0:
                return Response({
                    'error': 'No valid data found in the file. Please check the file format.',
                    'expected_columns': [
                        'Code_Enseignant', 'nom', 'prenom', 'nom_jeune_fille', 'genre', 
                        'etat', 'departement', 'grade', 'email1', 'email2', 'tel1', 'tel2'
                    ],
                    'note': 'First row should be headers, data starts from row 2. Supported formats: .xlsx, .xls, .csv'
                }, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'message': f'Successfully processed file',
                'inserted': inserted,
                'skipped': skipped,
                'summary': {
                    'total_rows': len(inserted) + len(skipped),
                    'inserted_count': len(inserted),
                    'skipped_count': len(skipped)
                }
            }, status=status.HTTP_201_CREATED)

        except BadZipFile:
            return Response({
                'error': 'Invalid Excel file. The file appears to be corrupted or is not a valid Excel file.',
                'hint': 'Please ensure you are uploading a valid .xlsx, .xls, or .csv file. If you saved a CSV as .xlsx, please upload it as .csv instead.'
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({
                'error': f'Error processing file: {str(e)}',
                'type': type(e).__name__,
                'hint': 'Please ensure your file has the correct format and column structure'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator

@method_decorator(csrf_exempt, name='dispatch')
class UploadFormations_xlsx(APIView):
    parser_classes = (MultiPartParser, FormParser)
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        # Log file details for debugging
        print(f"Received file: {excel_file.name}")
        print(f"File size: {excel_file.size} bytes")
        print(f"Content type: {excel_file.content_type}")

        try:
            # Check file extension
            file_name = excel_file.name.lower()
            
            if file_name.endswith('.csv'):
                # Handle CSV files
                import csv
                import io
                
                # Read CSV file
                decoded_file = excel_file.read().decode('utf-8-sig')  # utf-8-sig handles BOM
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string)
                
                rows = list(reader)
                if len(rows) < 2:
                    return Response({
                        'error': 'CSV file appears to be empty or only contains headers'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                inserted = []
                skipped = []
                
                # Skip header row
                for row_num, row in enumerate(rows[1:], start=2):
                    if len(row) < 8:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Insufficient columns. Expected at least 8, got {len(row)}'
                        })
                        continue
                    
                    # Process row data
                    domaine = row[0].strip() if row[0] else ''
                    filiere = row[1].strip() if len(row) > 1 else ''
                    niveau_cycle = row[2].strip() if len(row) > 2 else ''
                    specialites = row[3].strip() if len(row) > 3 else ''
                    nbr_sections = row[4].strip() if len(row) > 4 else ''
                    nbr_groupes = row[5].strip() if len(row) > 5 else ''
                    semestre = row[6].strip() if len(row) > 6 else ''
                    modules = row[7].strip() if len(row) > 7 else ''
                    
                    # Convert numeric fields
                    try:
                        nbr_sections_int = int(nbr_sections) if nbr_sections else None
                    except ValueError:
                        nbr_sections_int = None
                        
                    try:
                        nbr_groupes_int = int(nbr_groupes) if nbr_groupes else None
                    except ValueError:
                        nbr_groupes_int = None

                    # Check for duplicates
                    exists = Formations.objects.filter(
                        domaine=domaine,
                            filière=filiere,
                            niveau_cycle=niveau_cycle,
                            specialités=specialites,
                            nbr_sections=nbr_sections_int,
                            nbr_groupes=nbr_groupes_int,
                            semestre=semestre,
                            modules=modules
                    ).exists()
                    
                    if exists:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': 'Formation already exists with same domaine, filière, niveau_cycle, and specialités'
                        })
                        continue

                    try:
                        formation = Formations(
                            domaine=domaine,
                            filière=filiere,
                            niveau_cycle=niveau_cycle,
                            specialités=specialites,
                            nbr_sections=nbr_sections_int,
                            nbr_groupes=nbr_groupes_int,
                            semestre=semestre,
                            modules=modules
                        )
                        formation.save()
                        
                        inserted.append({
                            'id': formation.id,
                            'domaine': formation.domaine,
                            'filière': formation.filière,
                            'niveau_cycle': formation.niveau_cycle,
                            'specialités': formation.specialités
                        })
                    except Exception as e:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Database error: {str(e)}'
                        })
                
            else:
                # Handle .xlsx files with openpyxl
                workbook = load_workbook(filename=excel_file, data_only=True)
                sheet = workbook.active

                inserted = []
                skipped = []
                
                # Check if sheet has any data
                if sheet.max_row < 2:
                    return Response({
                        'error': 'Excel file appears to be empty or only contains headers'
                    }, status=status.HTTP_400_BAD_REQUEST)

                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    # Skip empty rows
                    if not row or all(cell is None for cell in row):
                        continue
                        
                    # Check if row has at least 8 columns
                    if len(row) < 8:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Insufficient columns. Expected at least 8, got {len(row)}'
                        })
                        continue

                    # Extract values
                    domaine = str(row[0]).strip() if row[0] else ''
                    filiere = str(row[1]).strip() if row[1] else ''
                    niveau_cycle = str(row[2]).strip() if row[2] else ''
                    specialites = str(row[3]).strip() if row[3] else ''
                    nbr_sections = row[4] if row[4] is not None else None
                    nbr_groupes = row[5] if row[5] is not None else None
                    semestre = str(row[6]).strip() if row[6] else ''
                    modules = str(row[7]).strip() if row[7] else ''
                    
                    # Convert numeric fields
                    try:
                        if nbr_sections is not None:
                            nbr_sections_int = int(float(str(nbr_sections)))
                        else:
                            nbr_sections_int = None
                    except (ValueError, TypeError):
                        nbr_sections_int = None
                        
                    try:
                        if nbr_groupes is not None:
                            nbr_groupes_int = int(float(str(nbr_groupes)))
                        else:
                            nbr_groupes_int = None
                    except (ValueError, TypeError):
                        nbr_groupes_int = None

                    # Check for duplicates
                    exists = Formations.objects.filter(
                        domaine=domaine,
                        filière=filiere,
                        niveau_cycle=niveau_cycle,
                        specialités=specialites
                    ).exists()
                    
                    if exists:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': 'Formation already exists with same domaine, filière, niveau_cycle, and specialités'
                        })
                        continue

                    try:
                        formation = Formations(
                            domaine=domaine,
                            filière=filiere,
                            niveau_cycle=niveau_cycle,
                            specialités=specialites,
                            nbr_sections=nbr_sections_int,
                            nbr_groupes=nbr_groupes_int,
                            semestre=semestre,
                            modules=modules
                        )
                        formation.save()
                        
                        inserted.append({
                            'id': formation.id,
                            'domaine': formation.domaine,
                            'filière': formation.filière,
                            'niveau_cycle': formation.niveau_cycle,
                            'specialités': formation.specialités
                        })
                        
                    except Exception as e:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Database error: {str(e)}'
                        })

            # If nothing was processed, the file might be in wrong format
            if len(inserted) == 0 and len(skipped) == 0:
                return Response({
                    'error': 'No valid data found in the file. Please check the file format.',
                    'expected_columns': [
                        'domaine', 'filiere', 'niveau_cycle', 'specialites', 
                        'nbr_sections', 'nbr_groupes', 'semestre', 'modules'
                    ],
                    'note': 'First row should be headers, data starts from row 2. Supported formats: .xlsx, .csv'
                }, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'message': f'Successfully processed file',
                'inserted': inserted,
                'skipped': skipped,
                'summary': {
                    'total_rows': len(inserted) + len(skipped),
                    'inserted_count': len(inserted),
                    'skipped_count': len(skipped)
                }
            }, status=status.HTTP_201_CREATED)

        except BadZipFile:
            return Response({
                'error': 'Invalid Excel file. The file appears to be corrupted or is not a valid Excel file.',
                'hint': 'Please ensure you are uploading a valid .xlsx or .csv file.'
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({
                'error': f'Error processing file: {str(e)}',
                'type': type(e).__name__,
                'hint': 'Please ensure your file has the correct format and column structure'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)