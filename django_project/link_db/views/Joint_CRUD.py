
from zipfile import BadZipFile
from django.http import JsonResponse
from django.views.decorators.http import require_GET
from rest_framework.views import APIView
from django.views.decorators.csrf import csrf_exempt
from django.db import transaction
from requests import Response
from rest_framework_simplejwt.authentication import JWTAuthentication
from rest_framework.permissions import IsAuthenticated
from rest_framework import status
from link_db.models import Planning, Surveillant, Formations, Creneau, Enseignants
from rest_framework.parsers import MultiPartParser, FormParser
from openpyxl import load_workbook

import json
@require_GET
def get_planning_with_creneau_and_formation(request):
    """
    GET /api/planning-join/
    Retourne la liste de tous les plannings avec leur créneau et leur formation associée.
    """
    # On profite de select_related pour faire la jointure SQL sur formation et creneau
    queryset = Planning.objects.select_related('formation', 'id_creneau').all()

    data = []
    for plan in queryset:
        c = plan.id_creneau
        f = plan.formation

        data.append({
            'id_planning': plan.id_planning,
            'section': plan.section,
            'session': plan.session,
            'nombre_surveillant': plan.nombre_surveillant,
            'creneau': {
                'id_creneau': c.id_creneau,
                'date_creneau': c.date_creneau,
                'heure_creneau': c.heure_creneau,
                'salle': c.salle,
            },
            'formation': {
                'id': f.pk,
                'domaine': f.domaine,
                'filière': f.filière,
                'niveau_cycle': f.niveau_cycle,
                'specialités': f.specialités,
                'nbr_sections': f.nbr_sections,
                'nbr_groupes': f.nbr_groupes,
                'semestre': f.semestre,
                'modules': f.modules,
            }
        })

    return JsonResponse(data, safe=False)



@csrf_exempt
def delete_planning_only(request):
    if request.method != 'DELETE':
        return JsonResponse({'error': 'Only DELETE allowed'}, status=405)

    try:
        payload = json.loads(request.body)
        planning_id = payload.get('id_planning')
        if not planning_id:
            return JsonResponse({'error': 'Field "id_planning" is required'}, status=400)

        planning = Planning.objects.get(id_planning=planning_id)
        planning.delete()
        return JsonResponse({'message': f'Planning {planning_id} deleted successfully'}, status=200)

    except Planning.DoesNotExist:
        return JsonResponse({'error': f'No Planning found with id_planning = {planning_id}'}, status=404)

    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def create_planning_with_surveillants(request):
    """
    POST /api/create_planning_with_surveillants/
    Crée un Planning et plusieurs Surveillants liés.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST allowed'}, status=405)

    try:
        data = json.loads(request.body)
        
        # Log received data
        print(f"Create Planning - Received data: {json.dumps(data, indent=2)}")
        
        required = ('formation_id', 'section', 'nombre_surveillant', 'id_creneau', 'surveillants')
        for f in required:
            if f not in data:
                return JsonResponse({'error': f'Field "{f}" is required'}, status=400)

        surveillants = data['surveillants']
        if not isinstance(surveillants, list) or not surveillants:
            return JsonResponse({'error': '"surveillants" must be a non-empty list'}, status=400)

        # Verify and normalize est_charge_cours flag
        pcs = [s for s in surveillants if s.get('est_charge_cours') == 1]
        if len(pcs) > 1:
            return JsonResponse({'error': 'Only one surveillant can have est_charge_cours=1'}, status=400)
        if len(pcs) == 0:
            surveillants[0]['est_charge_cours'] = 1

        # Verify all entities exist before creating
        try:
            formation = Formations.objects.get(pk=data['formation_id'])
        except Formations.DoesNotExist:
            return JsonResponse({'error': f'Formation with id {data["formation_id"]} not found'}, status=404)

        try:
            creneau = Creneau.objects.get(pk=data['id_creneau'])
        except Creneau.DoesNotExist:
            return JsonResponse({'error': f'Creneau with id {data["id_creneau"]} not found'}, status=404)

        # Verify all enseignants exist
        for s in surveillants:
            code = s.get('code_enseignant')
            if not code:
                return JsonResponse({'error': 'Each surveillant needs "code_enseignant"'}, status=400)
            if not Enseignants.objects.filter(Code_Enseignant=code).exists():
                return JsonResponse({'error': f'Enseignant with code {code} not found'}, status=404)

        with transaction.atomic():
            # Create planning
            planning = Planning.objects.create(
                formation=formation,
                section=data['section'],
                nombre_surveillant=data['nombre_surveillant'],
                session=data.get('session', ''),
                id_creneau=creneau
            )

            created = []
            for s in surveillants:
                code = s.get('code_enseignant')
                enseignant = Enseignants.objects.get(Code_Enseignant=code)
                
                srv = Surveillant.objects.create(
                    id_planning=planning,
                    code_enseignant=enseignant,
                    est_charge_cours=s.get('est_charge_cours', 0)
                )
                created.append({
                    'id_surveillance': srv.id_surveillance,
                    'code_enseignant': code,
                    'est_charge_cours': srv.est_charge_cours
                })

        return JsonResponse({
            'message': 'Planning et surveillants créés avec succès',
            'id_planning': planning.id_planning,
            'surveillants': created
        }, status=201)

    except json.JSONDecodeError as e:
        return JsonResponse({'error': f'Invalid JSON: {str(e)}'}, status=400)
    except Exception as e:
        import traceback
        print(f"Error in create_planning_with_surveillants: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'error': f'Internal server error: {str(e)}'}, status=500)


@csrf_exempt   
def update_planning_with_surveillants(request):
    """
    PUT /api/update_planning_with_surveillants/
    Met à jour un Planning et son ensemble de Surveillants associés.
    """
    if request.method != 'PUT':
        return JsonResponse({'error': 'Only PUT allowed'}, status=405)

    try:
        data = json.loads(request.body)
        
        # Log received data
        print(f"Update Planning - Received data: {json.dumps(data, indent=2)}")
        
        planning_id = data.get('id_planning')
        surveillants = data.get('surveillants')

        if not planning_id:
            return JsonResponse({'error': '"id_planning" is required'}, status=400)
        if not isinstance(surveillants, list) or not surveillants:
            return JsonResponse({'error': '"surveillants" must be a non-empty list'}, status=400)

        # Vérification du flag est_charge_cours
        principals = [s for s in surveillants if s.get('est_charge_cours') == 1]
        if len(principals) > 1:
            return JsonResponse({'error': 'Only one surveillant can have est_charge_cours=1'}, status=400)
        if len(principals) == 0:
            surveillants[0]['est_charge_cours'] = 1

        with transaction.atomic():
            # Get the existing planning
            try:
                planning = Planning.objects.get(id_planning=planning_id)
            except Planning.DoesNotExist:
                return JsonResponse({'error': f'Planning with id {planning_id} not found'}, status=404)

            # Update planning fields if provided
            if 'formation_id' in data:
                try:
                    formation = Formations.objects.get(pk=data['formation_id'])
                    planning.formation = formation
                except Formations.DoesNotExist:
                    return JsonResponse({'error': f'Formation with id {data["formation_id"]} not found'}, status=404)

            if 'section' in data:
                planning.section = data['section']
            
            if 'nombre_surveillant' in data:
                planning.nombre_surveillant = data['nombre_surveillant']
            
            if 'session' in data:
                planning.session = data['session']
            
            if 'id_creneau' in data:
                try:
                    creneau = Creneau.objects.get(pk=data['id_creneau'])
                    planning.id_creneau = creneau
                except Creneau.DoesNotExist:
                    return JsonResponse({'error': f'Creneau with id {data["id_creneau"]} not found'}, status=404)

            # Save the updated planning
            planning.save()

            # Delete existing surveillants and create new ones
            Surveillant.objects.filter(id_planning=planning).delete()

            created = []
            for s in surveillants:
                code = s.get('code_enseignant')
                if not code:
                    transaction.set_rollback(True)
                    return JsonResponse({'error': 'Each surveillant needs "code_enseignant"'}, status=400)
                
                # Verify if enseignant exists
                try:
                    enseignant = Enseignants.objects.get(Code_Enseignant=code)
                except Enseignants.DoesNotExist:
                    transaction.set_rollback(True)
                    return JsonResponse({'error': f'Enseignant with code {code} not found'}, status=404)

                srv = Surveillant.objects.create(
                    id_planning=planning,
                    code_enseignant=enseignant,
                    est_charge_cours=s.get('est_charge_cours', 0)
                )
                created.append({
                    'id_surveillance': srv.id_surveillance,
                    'code_enseignant': code,
                    'est_charge_cours': srv.est_charge_cours
                })

        # Return the updated planning data with IDs (not objects)
        return JsonResponse({
            'message': 'Planning et surveillants mis à jour avec succès',
            'planning': {
                'id_planning': planning.id_planning,
                'formation_id': planning.formation.pk if planning.formation else None,
                'section': planning.section,
                'nombre_surveillant': planning.nombre_surveillant,
                'session': planning.session,
                'id_creneau': planning.id_creneau.pk if planning.id_creneau else None
            },
            'surveillants': created
        }, status=200)

    except json.JSONDecodeError as e:
        return JsonResponse({'error': f'Invalid JSON: {str(e)}'}, status=400)
    except Exception as e:
        import traceback
        print(f"Error in update_planning_with_surveillants: {str(e)}")
        print(traceback.format_exc())
        return JsonResponse({'error': f'Internal server error: {str(e)}'}, status=500)
    

@csrf_exempt
def get_surveillants_by_planning(request):
    """Get all surveillants for a specific planning ID"""
    if request.method != 'GET':
        return JsonResponse({'error': 'Only GET requests are allowed'}, status=405)
    
    try:
        id_planning = request.GET.get('id_planning')
        
        if not id_planning:
            return JsonResponse({'error': 'id_planning parameter is required'}, status=400)
        
        # Get surveillants with related enseignant data
        surveillants = Surveillant.objects.filter(
            id_planning_id=id_planning
        ).select_related('code_enseignant')
        
        result = []
        for surveillant in surveillants:
            enseignant = surveillant.code_enseignant
            
            # Build the response carefully, avoiding the problematic field
            surveillant_data = {
                'id_surveillance': surveillant.id_surveillance,
                'id_planning': surveillant.id_planning_id,
                'code_enseignant': enseignant.Code_Enseignant,
                'est_charge_cours': surveillant.est_charge_cours,
                'enseignant': {
                    'Code_Enseignant': enseignant.Code_Enseignant,
                    'nom': enseignant.nom or '',
                    'prenom': enseignant.prenom or '',
                    'nom_jeune_fille': enseignant.nom_jeune_fille or '',
                    'genre': enseignant.genre or '',
                    'etat': enseignant.etat or '',
                    'grade': enseignant.grade or '',
                    'email1': enseignant.email1 or '',
                    'email2': enseignant.email2 or '',
                    'tel1': enseignant.tel1 or '',
                    'tel2': enseignant.tel2 or '',
                    # Handle département field carefully
                    # Try to get the value without directly accessing the field
                    'département': getattr(enseignant, 'département', '') if hasattr(enseignant, 'département') else ''
                }
            }
            result.append(surveillant_data)
        
        return JsonResponse(result, safe=False, json_dumps_params={'ensure_ascii': False})
        
    except Exception as e:
        import logging
        logger = logging.getLogger(__name__)
        logger.exception("Error in get_surveillants_by_planning")  
@csrf_exempt
def get_monitoring_planning(request):
    monitoring_data = []
    
    try:
        # Get all plannings with related data
        plannings = Planning.objects.select_related('id_creneau', 'formation').prefetch_related('surveillant_set__code_enseignant')
        
        for planning in plannings:
            # Get all surveillants for this planning
            for surveillant in planning.surveillant_set.all():
                enseignant = surveillant.code_enseignant
                
                monitoring_data.append({
                    'teacher_name': f"{enseignant.prenom} {enseignant.nom}".strip() if enseignant.prenom and enseignant.nom else enseignant.Code_Enseignant,
                    'teacher_code': enseignant.Code_Enseignant,
                    'module': planning.formation.modules,
                    'room': planning.id_creneau.salle,
                    'date': str(planning.id_creneau.date_creneau),
                    'time': str(planning.id_creneau.heure_creneau),
                    'level': planning.formation.niveau_cycle,
                    'specialty': planning.formation.filière,
                    'role': 'Main' if surveillant.est_charge_cours == 1 else 'Assistant'
                })
        
        # Sort by teacher name, then by date and time
        monitoring_data.sort(key=lambda x: (x['teacher_name'], x['date'], x['time']))
        
        return JsonResponse(monitoring_data, safe=False)
        
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    

# Add this class to your Xcel_Api.py file

class UploadFormations_xlsx(APIView):
    parser_classes = (MultiPartParser, FormParser)
    authentication_classes = [JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            return Response({'error': 'No file uploaded'}, status=status.HTTP_400_BAD_REQUEST)

        # Log file details for debugging
        print(f"Received file: {excel_file.name}")
        print(f"File size: {excel_file.size} bytes")
        print(f"Content type: {excel_file.content_type}")

        try:
            # Check file extension
            file_name = excel_file.name.lower()
            
            if file_name.endswith('.csv'):
                # Handle CSV files
                import csv
                import io
                
                # Read CSV file
                decoded_file = excel_file.read().decode('utf-8-sig')  # utf-8-sig handles BOM
                io_string = io.StringIO(decoded_file)
                reader = csv.reader(io_string)
                
                rows = list(reader)
                if len(rows) < 2:
                    return Response({
                        'error': 'CSV file appears to be empty or only contains headers'
                    }, status=status.HTTP_400_BAD_REQUEST)
                
                inserted = []
                skipped = []
                
                # Skip header row
                for row_num, row in enumerate(rows[1:], start=2):
                    if len(row) < 8:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Insufficient columns. Expected at least 8, got {len(row)}'
                        })
                        continue
                    
                    # Process row data
                    domaine = row[0].strip() if row[0] else ''
                    filiere = row[1].strip() if len(row) > 1 else ''
                    niveau_cycle = row[2].strip() if len(row) > 2 else ''
                    specialites = row[3].strip() if len(row) > 3 else ''
                    nbr_sections = row[4].strip() if len(row) > 4 else ''
                    nbr_groupes = row[5].strip() if len(row) > 5 else ''
                    semestre = row[6].strip() if len(row) > 6 else ''
                    modules = row[7].strip() if len(row) > 7 else ''
                    
                    # Convert numeric fields
                    try:
                        nbr_sections_int = int(nbr_sections) if nbr_sections else None
                    except ValueError:
                        nbr_sections_int = None
                        
                    try:
                        nbr_groupes_int = int(nbr_groupes) if nbr_groupes else None
                    except ValueError:
                        nbr_groupes_int = None

                    # Check for duplicates
                    exists = Formations.objects.filter(
                        domaine=domaine,
                        filière=filiere,
                        niveau_cycle=niveau_cycle,
                        specialités=specialites
                    ).exists()
                    
                    if exists:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': 'Formation already exists with same domaine, filière, niveau_cycle, and specialités'
                        })
                        continue

                    try:
                        formation = Formations(
                            domaine=domaine,
                            filière=filiere,
                            niveau_cycle=niveau_cycle,
                            specialités=specialites,
                            nbr_sections=nbr_sections_int,
                            nbr_groupes=nbr_groupes_int,
                            semestre=semestre,
                            modules=modules
                        )
                        formation.save()
                        
                        inserted.append({
                            'id': formation.id,
                            'domaine': formation.domaine,
                            'filière': formation.filière,
                            'niveau_cycle': formation.niveau_cycle,
                            'specialités': formation.specialités
                        })
                    except Exception as e:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Database error: {str(e)}'
                        })
                
            else:
                # Handle .xlsx files with openpyxl
                workbook = load_workbook(filename=excel_file, data_only=True)
                sheet = workbook.active

                inserted = []
                skipped = []
                
                # Check if sheet has any data
                if sheet.max_row < 2:
                    return Response({
                        'error': 'Excel file appears to be empty or only contains headers'
                    }, status=status.HTTP_400_BAD_REQUEST)

                for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    # Skip empty rows
                    if not row or all(cell is None for cell in row):
                        continue
                        
                    # Check if row has at least 8 columns
                    if len(row) < 8:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Insufficient columns. Expected at least 8, got {len(row)}'
                        })
                        continue

                    # Extract values
                    domaine = str(row[0]).strip() if row[0] else ''
                    filiere = str(row[1]).strip() if row[1] else ''
                    niveau_cycle = str(row[2]).strip() if row[2] else ''
                    specialites = str(row[3]).strip() if row[3] else ''
                    nbr_sections = row[4] if row[4] is not None else None
                    nbr_groupes = row[5] if row[5] is not None else None
                    semestre = str(row[6]).strip() if row[6] else ''
                    modules = str(row[7]).strip() if row[7] else ''
                    
                    # Convert numeric fields
                    try:
                        if nbr_sections is not None:
                            nbr_sections_int = int(float(str(nbr_sections)))
                        else:
                            nbr_sections_int = None
                    except (ValueError, TypeError):
                        nbr_sections_int = None
                        
                    try:
                        if nbr_groupes is not None:
                            nbr_groupes_int = int(float(str(nbr_groupes)))
                        else:
                            nbr_groupes_int = None
                    except (ValueError, TypeError):
                        nbr_groupes_int = None

                    # Check for duplicates
                    exists = Formations.objects.filter(
                        domaine=domaine,
                        filière=filiere,
                        niveau_cycle=niveau_cycle,
                        specialités=specialites
                    ).exists()
                    
                    if exists:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': 'Formation already exists with same domaine, filière, niveau_cycle, and specialités'
                        })
                        continue

                    try:
                        formation = Formations(
                            domaine=domaine,
                            filière=filiere,
                            niveau_cycle=niveau_cycle,
                            specialités=specialites,
                            nbr_sections=nbr_sections_int,
                            nbr_groupes=nbr_groupes_int,
                            semestre=semestre,
                            modules=modules
                        )
                        formation.save()
                        
                        inserted.append({
                            'id': formation.id,
                            'domaine': formation.domaine,
                            'filière': formation.filière,
                            'niveau_cycle': formation.niveau_cycle,
                            'specialités': formation.specialités
                        })
                        
                    except Exception as e:
                        skipped.append({
                            'row': f'Row {row_num}',
                            'reason': f'Database error: {str(e)}'
                        })

            # If nothing was processed, the file might be in wrong format
            if len(inserted) == 0 and len(skipped) == 0:
                return Response({
                    'error': 'No valid data found in the file. Please check the file format.',
                    'expected_columns': [
                        'domaine', 'filiere', 'niveau_cycle', 'specialites', 
                        'nbr_sections', 'nbr_groupes', 'semestre', 'modules'
                    ],
                    'note': 'First row should be headers, data starts from row 2. Supported formats: .xlsx, .csv'
                }, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                'message': f'Successfully processed file',
                'inserted': inserted,
                'skipped': skipped,
                'summary': {
                    'total_rows': len(inserted) + len(skipped),
                    'inserted_count': len(inserted),
                    'skipped_count': len(skipped)
                }
            }, status=status.HTTP_201_CREATED)

        except BadZipFile:
            return Response({
                'error': 'Invalid Excel file. The file appears to be corrupted or is not a valid Excel file.',
                'hint': 'Please ensure you are uploading a valid .xlsx or .csv file.'
            }, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            import traceback
            traceback.print_exc()
            return Response({
                'error': f'Error processing file: {str(e)}',
                'type': type(e).__name__,
                'hint': 'Please ensure your file has the correct format and column structure'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)    